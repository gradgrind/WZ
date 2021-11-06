# -*- coding: utf-8 -*-
"""
tables/spreadsheet.py

Last updated:  2021-10-21

Spreadsheet file reader, returning all cells as strings.
For reading, simple tsv files (no quoting, no escapes), Excel files (.xlsx)
and Open Document files (.ods) are supported.

Spreadsheet file writers, table contains only strings (and empty cells).
For writing, only simple unformatted xlsx files and tsv files
(no quoting, no escapes) are supported.

Dates are read and written as strings in the format 'yyyy-mm-dd'.

=+LICENCE=============================
Copyright 2021 Michael Towers

   Licensed under the Apache License, Version 2.0 (the "License");
   you may not use this file except in compliance with the License.
   You may obtain a copy of the License at

       http://www.apache.org/licenses/LICENSE-2.0

   Unless required by applicable law or agreed to in writing, software
   distributed under the License is distributed on an "AS IS" BASIS,
   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
   See the License for the specific language governing permissions and
   limitations under the License.

=-LICENCE========================================
"""

### Messages
_UNSUPPORTED_FILETYPE   = "Nicht unterstützer Dateityp ({ending})"
_TABLENOTFOUND          = "Tabellendatei existiert nicht:\n   {path}"
_MULTIPLEMATCHINGFILES  = "Mehrere passende Dateien:\n   {path}"
_TABLENOTREADABLE       = "Tabellendatei konnte nicht eingelesen werden:\n   {path}"
_INVALIDSHEETNAME       = "Ungültige Tabellenname: '{name}'"
_INVALIDCELLNAME        = "Ungültiger Zellenbezeichnung: '{name}'"
_INVALID_FILE           = "Ungültige oder fehlerhafte Datei"
_NO_TYPE_EXTENSION      = "Dateityp-Erweiterung fehlt: {fname}"
_DUPLICATE_COLUMN_NAME  = "Spaltenname doppelt vorhanden: {name}"
_ESSENTIAL_FIELD_MISSING = "Feld (Spalte) '{field}' fehlt in der Tabelle"
_ESSENTIAL_FIELD_EMPTY  = "Feld (Spalte) '{field}' darf nicht leer sein"
_ESSENTIAL_INFO_MISSING = "Info-Feld '{field}' fehlt in der Tabelle"
_ESSENTIAL_INFO_EMPTY   = "Info-Feld '{field}' darf nicht leer sein"
_INFO_IN_BODY           = "Info-Zeilen müssen vor der Kopfzeile stehen"

########################################################################

import sys, os, datetime, re
if __name__ == '__main__':
    # Enable package import if running as module
    this = sys.path[0]
    appdir = os.path.dirname(this)
    sys.path[0] = appdir
    basedir = os.path.dirname(appdir)

### +++++

import io

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

from tables.simple_ods_reader import OdsReader

NONE = ''

### -----

class TsvReader(dict):
    def __init__ (self, filepath):
        """Read a tab-separated-value table as a list of rows,
        each row is a list of cell values.
        <filepath> can be the path to a tsv file, but it could be an
        <io.BytesIO> object.
        This format doesn't support multiple sheets, the single table
        is named 'TSV' and the resulting instance has only one key, 'TSV'.
        All values are returned as "stripped" strings, except for empty
        cells, these having the value <None>.
        """
        super().__init__()
        if type(filepath) == str:
            with open(filepath, 'rb') as fbi:
                lines = fbi.read().splitlines()
        else:
            lines = filepath.read().splitlines()
        rows = []
        maxlen = 0
        for row_b in lines:
            #print(repr(row_b))
            row = [cell.decode('utf-8').strip() or None
                    for cell in row_b.split(b'\t')]
            l = len(row)
            if l > maxlen:
                maxlen = l
            rows.append(row)
        for row in rows:
            dl = maxlen - len(row)
            if dl:
                row += [None] * dl
        self['TSV'] = rows
#
    def mergedRanges (self, sheetname):
        """Returns an empty list as tsv doesn't support cell merging.
        """
        return []

###

class XlsReader(dict):
    def __init__ (self, filepath):
        """Read an Excel spreadsheet as a list of rows,
        each row is a list of cell values.
        All sheets in the file are read, a <dict> of sheets
        (name -> row list) is built.

        This is a read-only utility. Formulae, style, etc. are not retained.
        For formulae the last-calculated value is returned.
        All values are returned as strings.
        """
        super().__init__()
        self._mergedRanges = {}
        # Note that <data_only=True> replaces all formulae by their value,
        # which is probably good for reading, but not for writing!
        wb = load_workbook(filepath, data_only=True)
        for wsname in wb.sheetnames:
            ws = wb[wsname]
            rows = []
            for row in ws.iter_rows():
                values = []
                for cell in row:
                    v = cell.value
                    if type(v) == datetime.datetime:
                        v = v.strftime("%Y-%m-%d")
                    elif type(v) == str:
                        v = v.strip()
                        if v == NONE:
                             v = None
                    elif v != None:
                        v = str(v)
                    values.append(v)
                rows.append(values)
            self[wsname] = rows
            self._mergedRanges[wsname] = ws.merged_cells.ranges
#
    def mergedRanges(self, sheetname):
        """Returns a list like ['AK2:AM2', 'H33:AD33', 'I34:J34', 'L34:AI34'].
        """
        return self._mergedRanges[sheetname]

###

class TableError(Exception):
    pass

###

def spreadsheet_file_complete(filepath):
    """Determine the file-type extension if it is missing.
    Check that it is one of the supported table formats.
    Return the full path including file-type extension.
    """
    f_e = filepath.rsplit('.', 1)
    if len(f_e) == 2:
        if f_e[1] in Spreadsheet._SUPPORTED_TYPES:
            if os.path.isfile(filepath):
                return filepath
            raise TableError(_TABLENOTFOUND.format(
                    path = filepath))
    # No type-extension provided, test valid possibilities
    found = None
    for x in Spreadsheet._SUPPORTED_TYPES:
        fp = f'{filepath}.{x}'
        if os.path.isfile(fp):
            if found:
                raise TableError(_MULTIPLEMATCHINGFILES.format(
                        path = fpbase))
            found = fp
    if found:
        return found
    raise TableError(_TABLENOTFOUND.format(
            path = filepath))
##
class Spreadsheet:
    """This class manages a (read-only) representation of a spreadsheet file.
    The individual table/sheet names are available via the method
    <getTableNames()>.
    The currently selected table can be set using the method
    <setTable('sheetname')>.
    The first table is accessed by default. To access others, an optional
    argument must be passed.
    Row length and column length are available via the methods rowLen()
    and colLen().
    The value of a cell is read using <getValue(row, col)>, where <row>
    and <col> are 0-based indexes.
    All cell values are strings, or <None> if empty.
    """
    _SUPPORTED_TYPES = {'tsv': TsvReader, 'ods': OdsReader, 'xlsx': XlsReader}
    @classmethod
    def filetype_endings(cls):
        return list(cls._SUPPORTED_TYPES)
    #+
    @classmethod
    def supportedType(cls, filename):
        """Check the ending of a file name (or path).
        Return <True> if the type is supported, else <False>.
        """
        fsplit = filename.rsplit('.', 1)
        if len(fsplit) == 2:
            if fsplit[1] in cls._SUPPORTED_TYPES:
                return True
        return False
#
    def __init__(self, filepath):
        """The filepath can be passed with or without type-extension.
        If no type-extension is given, the folder will be searched for a
        suitable file.
        Alternatively, <filepath> may be an in-memory binary stream
        (io.BytesIO) with attribute 'filename' (so that the
        type-extension can be read).
        """
        self._spreadsheet = None
        self._sheetNames = None
        self._table = None
        self.ixHeaderEnd = None
        if type(filepath) == str:
            # realfile = True
            filepath = spreadsheet_file_complete(filepath)
            self.filename = os.path.basename(filepath)
            self.filepath = filepath
            ending = self.filename.rsplit('.', 1)[1]
        else:
            # realfile = False
            try:
                self.filename = filepath.filename
            except:
                raise TableError(_INVALID_FILE)
            try:
                ending = self.filename.rsplit('.', 1)[1]
            except:
                raise TableError(_NO_TYPE_EXTENSION.format(
                        fname=self.filename))
            self.filepath = None
        try:
            handler = self._SUPPORTED_TYPES [ending]
        except:
            raise TableError(_UNSUPPORTED_FILETYPE.format(ending=ending))
        try:
            self._spreadsheet = handler(filepath)
        except:
            raise TableError(_TABLENOTREADABLE.format(
                    path=self.filepath or self.filename))
        self._sheetNames = list(self._spreadsheet)
        # Default sheet is the first:
        self._table = self._spreadsheet[self._sheetNames[0]]
#
    def table(self):
        """Return the current table (initially the first sheet).
        """
        return self._table
#
    def rowLen(self, table = None):
        if not table:
            table = self._table
        return len(table[0])
#
    def colLen(self, table = None):
        if not table:
            table = self._table
        return len(table)
#
    def getValue(self, rx, cx, table = None):
        if not table:
            table = self._table
        return table[rx][cx]
#
    def getABValue(self, A1, table = None):
        r, c = self.rowcol(A1)
        return self.getValue(r, c, table)
#
    def getTableNames(self):
        return self._sheetNames
#
    def _getTable(self, tablename, failerror = True):
        try:
            #print (self._spreadsheet.keys())
            return self._spreadsheet[tablename]
        except:
            if failerror:
                raise TableError(_INVALIDSHEETNAME.format(name=tablename))
            return None
#
    def setTable(self, tablename):
        table = self._getTable(tablename)
        if table:
            self._table = table
            return True
        else:
            return False
#
#TODO: Is this needed?
    def getColumnHeaders(self, rowix, table = None):
        """Return a dict of table headers, header -> column index.
        The row containing the headers is passed as argument.
        """
        self.ixHeaderEnd = None
        headers = {}
        for cellix in range(self.rowLen(table)):
            cellV = self.getValue(rowix, cellix, table)
            if cellV:
                if cellV == '#':
                    continue
                if cellV == '!':
                    self.ixHeaderEnd = cellix
                    break
                headers[cellV] = cellix
        return headers
#
#TODO: Is this needed?
    def getMergedRanges(self, tablename):
        return self._spreadsheet.mergedRanges(tablename)
#
    @staticmethod
    def rowcol(cellname):
        """Return a tuple (row, column) representing a cell position from
        the given reference in the spreadsheet form, e.g. "B12".
        """
        cell = cellname.upper()
        col = -1
        baseval = ord("A")
        i = 0
        for c in cell:
            v = ord(c)
            if v < baseval:
                break
            i += 1
            col = (col + 1) * 26 + v - baseval
        try:
            assert col >= 0
            return(int (cell[i:]) - 1, col)
        except:
            raise TableError(_INVALIDCELLNAME.format(name=cellname))
#
    @staticmethod
    def cellname(row, col):
        """Return the name of a cell given its coordinates (0-based):
        """
        return get_column_letter(col+1) + str(row+1)

###

def read_DataTable_filetypes(): return tuple(Spreadsheet._SUPPORTED_TYPES)

def read_DataTable(filepath_or_stream):
    """<filepath_or_stream> is a full file-path or in-memory stream as
    for <Spreadsheet>, which is used to read the file.
    Read a "DataTable", which is a specially structured table designed
    for storing lines of records (like in a relational database), but
    also having a header section for key-value entries relevant for
    the whole table.
    Return the table data as a mapping:
        {   '__INFO__': {key: value, ... },
            '__FIELDS__': [field, ... ],
            '__ROWS__': [{field: value, ... }, ... ]
        }

    Lines whose first field (column 0) is empty are ignored.
    Only string data is supported. Any fields in the input which are not
    intrinsically strings are converted (this is done by the
    <Spreadsheet> class). Also empty cells are strings ('').

    The first rows are set aside for group information. This is a list
    of key-value pairs. The first column contains '+++', the second
    the key, the third the value. Subsequent columns are ignored.
    These pairs are returned as the '__INFO__' value.

    After these, the first row with an entry in the first column is the
    header row, it contains the field names. Columns with no entry in
    this line will be excluded from the data (ignored). A list of field
    names is returned as the '__FIELDS__' value.

    The records are returned as a list of mappings {field: value}. This
    list is available as the '__ROWS__' value.
    """
    table = Spreadsheet(filepath_or_stream).table()
    rows = []
    info = {}
    header = []
    fields = []
    for row in table:
        c1 = row[0]
        if not c1:
            continue
        if header:
            # The header line has already been found.
            rowmap = {}
            for f, i in header:
                # <Spreadsheet> can return <None> in a cell:
                rowmap[f] = row[i] or NONE
            rows.append(rowmap)
        elif c1 == '+++':
            if header:
                raise TableError(_INFO_IN_BODY)
            info[row[1]] = row[2] or NONE
        else:
            # The field names
            i = 0   # column indexing
            header = []
            for f in row:
                if f:
                    if f in header:
                        raise TableError(_DUPLICATE_COLUMN_NAME.format(
                                name = f))
                    header.append((f, i))
                    fields.append(f)
                i += 1
    return {    '__INFO__': info,
                '__FIELDS__': fields,
                '__ROWS__': rows
    }
#
def filter_DataTable(data, fieldlist, infolist, extend = True):
    """Process the table data into mappings based on the two
    lists, <fields> and <info>, allowing translation of the field/key
    names to internal versions. Only those fields which are in these
    lists will be retained.
    Empty fields are guaranteed to contain ''.
    If <extend> is true, fields which are in the lists but not in
    the table will be added (though empty).
    <fieldlist> and <infolist> are lists of triples (or longer tuples):
        [   internal field-name,
            external (translated) field-name ( or false, e.g. ''),
            essential field (true/false),
            ... (possible further entries)
        ]
    If the external field name evaluates "false" (''), the internal
    and external names are identical.
    Return a mapping with the entries, '__INFO__' (the info-mapping),
    '__FIELDS__' (the list of internal field names) and '__ROWS__'
    (the list of row mappings).
    """
    tinfo = data['__INFO__']
    newinfo = {}
    for f, t, needed, *x in infolist:
        name = t or f   # null <t> => no translation, use internal name
        try:
            val = tinfo[name]
            if needed and not val:
                raise TableError(_ESSENTIAL_INFO_EMPTY.format(
                        field = name))
        except KeyError:
            if needed:
                raise TableError(_ESSENTIAL_INFO_MISSING.format(
                        field = name))
            if extend:
                val = NONE
        newinfo[f] = val
    # Check available fields against desired fields
    tfields = set(data['__FIELDS__'])
    fieldnames = []
    flist = []
    fname = {}
    for ftn in fieldlist:
        name = ftn[1] or ftn[0]   # null <t> => no translation, ...
        if name in tfields:
            flist.append(ftn)
            fieldnames.append(ftn[0])
        elif ftn[2]:
            raise TableError(_ESSENTIAL_FIELD_MISSING.format(
                            field = name))
        elif extend:
            flist.append(ftn)
            fieldnames.append(ftn[0])
        fname[ftn[0]] = name
    # Add the data rows
    rowmaps = []
    for row in data['__ROWS__']:
        rowmap = {}
        rowmaps.append(rowmap)
        for f, t, needed, *x in flist:
            name = t or f # null <t> => no translation, use internal name
            val = row.get(name)
            if needed and not val:
                raise TableError(_ESSENTIAL_FIELD_EMPTY.format(
                        field = fname[f]))
            rowmap[f] = val or NONE
    return {    '__INFO__': newinfo,
                '__FIELDS__': fieldnames,
                '__ROWS__': rowmaps
    }

###

def make_DataTable_filetypes(): return ('tsv', 'xlsx')

def make_DataTable(data, filetype,
        fieldlist = None, infolist = None,
        extend = True,  **xinfo):
    """Build a DataTable with info-lines, header-line and records.
    <data> is a mapping as returned by <read_DataTable>.
    <filetype> specifies which of the file-types in
    <make_DataTable_filetypes()> is to be generated.
    <fieldlist> is a list of triples (or longer tuples):
        [[internal-name, external-name, necessary, ... ], ... ]
        "necessary" is true if the field must be present and not
        empty in the supplied data.
    If <fieldlist> is not supplied, use the fields in the provided
    data.
    <infolist> is a similar list of triples, but for the info-lines.
    If <extend> is true, fields in <infolist> or <fieldlist> which
    are not supplied in the data will be added (does not apply to
    "necessary" fields, because if one of these is missing the
    function will fail).
    <xinfo> optionallyadds info fields.
    The file is returned as a <bytes> object.
    """
    if filetype == 'xlsx':
        table = NewSpreadsheet()
    elif filetype == 'tsv':
        table = NewTable()
    else:
        raise TableError(_UNSUPPORTED_FILETYPE.format(ending = filetype))
    hasinfo = 0
    info = data['__INFO__']
    info.update(xinfo)
    if infolist:
        for f, t, needed, *x in infolist:
            try:
                val = info[f]
                if needed and not val:
                    raise TableError(_ESSENTIAL_INFO_EMPTY.format(
                            field = f))
                # null <t> => no translation, use internal name
                table.add_row(('+++', t or f, val or None))
                hasinfo += 1
            except KeyError:
                if needed:
                    raise TableError(_ESSENTIAL_INFO_MISSING.format(
                            field = name))
                if extend:
                    # null <t> => no translation, use internal name
                    table.add_row(('+++', t or f, None))
                    hasinfo += 1
    else:
        for key, val in info.items():
            table.add_row(('+++', key, val))
            hasinfo += 1
    if hasinfo:
        table.add_row(None)
    tfields = data['__FIELDS__']
    if fieldlist:
        # Check available fields against desired fields
        fieldnames = []
        flist = []
        for f, t, needed, *x in fieldlist:
            if f in tfields:
                flist.append((f, needed))
                # null <t> => no translation, use internal name
                fieldnames.append(t or f)
            elif needed:
                raise TableError(_ESSENTIAL_FIELD_MISSING.format(
                                field = name))
            elif extend:
                flist.append((f, needed))
                fieldnames.append(t or f)
        table.add_row(fieldnames)
    else:
        flist = [(f, False) for f in tfields]
        table.add_row(tfields)
    # Add the data rows
    for row in data['__ROWS__']:
        rowvals = []
        for f, needed in flist:
            val = row.get(f)
            if needed and not val:
                raise TableError(_ESSENTIAL_FIELD_EMPTY.format(
                        field = f))
            rowvals.append(val or None)
        table.add_row(rowvals)
    return table.save()

###

class NewTable:
    """Build a tsv-table.
    The characters '\t', '\n' and '\r' are filtered out of the input
    strings.
    """
    def __init__(self):
        self._rowlist = []
#
    @staticmethod
    def _filter(text):
        return re.sub('\t\n\r', NONE, str(text)) if text else NONE
#
    def add_row(self, items):
        """Add a row with the values listed in <items>. The values will
        all be read as strings.
        """
        if items:
            self._rowlist.append('\t'.join([self._filter(item)
                    for item in items]))
        else:
            self._rowlist.append(NONE)
#
    def save(self, filepath = None):
        """If <filepath> is given, the resulting table will be written
        to a file. The ending '.tsv' is added automatically if it is
        not present already. Then return the full filepath.
        Without <filepath>, a <bytes> object is returned.
        """
        tbytes = '\n'.join(self._rowlist).encode('utf-8') + b'\n'
        if filepath:
            if not filepath.endswith('.tsv'):
                filepath += '.tsv'
            with open(filepath, 'wb') as fh:
                fh.write(tbytes)
            return filepath
        else:
            return tbytes

###

class NewSpreadsheet:
    """Build a simple xlsx table from scratch.
    """
    def __init__(self):
        # Create the workbook and worksheet we'll be working with
        self._wb = Workbook()
        self._ws = self._wb.active
        self._row = 0   # row counter, for serial row addition
#
    def set_cell(self, row, col, value):
        """Set a cell value (string only), using 0-based indexing.
        """
        self._ws.cell(row = row + 1, column = col + 1,
                value = NONE if value == None else str(value))
#
    def add_row(self, items):
        """Add a row with the values listed in <items>. The values will
        all be read as strings.
        """
        if items:
            col = 0
            for item in items:
                self.set_cell(self._row, col, item)
                col += 1
        self._row += 1
#
    def save(self, filepath = None):
        """If <filepath> is given, the resulting table will be written
        to a file. The ending '.xlsx' is added automatically if it is
        not present already. Then return the full filepath.
        Without <filepath>, a <bytes> object is returned.
        """
        if filepath:
            if not filepath.endswith('.xlsx'):
                filepath += '.xlsx'
            self._wb.save(filepath)
            return filepath
        else:
            virtual_workbook = io.BytesIO()
            self._wb.save(virtual_workbook)
            return virtual_workbook.getvalue()


#--#--#--#--#--#--#--#--#--#--#--#--#--#--#--#--#--#--#--#--#--#--#--#

if __name__ == '__main__':
    from core.base import start
    start.setup(os.path.join(basedir, 'TESTDATA'))

    for f in sys.argv[1:]:
        if f[0] != '-':
            dbt = read_DataTable(f)
            print("\nINFO:", dbt['__INFO__'])
            print("\nFIELDS:", dbt['__FIELDS__'])
            print("\nCONTENT:")
            for row in dbt['__ROWS__']:
                print(" :::", row)
            quit(0)

    filepath = DATAPATH('testing/Test1.tsv')
    fname = os.path.basename(filepath)
    tsv = TsvReader(filepath)
    print("\nROWS:")
    for row in tsv['TSV']:
        print(" :::", row)
    print("\n\nAnd now using a file-like object ...\n")
    with open(filepath, 'rb') as fbi:
        bytefile = fbi.read()
    flo = io.BytesIO(bytefile)
    flo.filename = fname
    tsv = TsvReader(flo)
    print("\nROWS:")
    for row in tsv['TSV']:
        print(" :::", row)

    dbt = read_DataTable(filepath)
    print("\nINFO:", dbt['__INFO__'])
    print("\nFIELDS:", dbt['__FIELDS__'])
    print("\nCONTENT:")
    for row in dbt['__ROWS__']:
        print(" :::", row)

    print("\nGRADES 10:")
    dbt = read_DataTable(DATAPATH('testing/Noten/NOTEN_2/Noten_10_2'))
    print("\nINFO:", dbt['__INFO__'])
    print("\nFIELDS:", dbt['__FIELDS__'])
    print("\nCONTENT:")
    for row in dbt['__ROWS__']:
        print(" :::", row)

    print("\nPUPILS + extend:")
    _dbt = read_DataTable(DATAPATH('testing/delta_test_pupils_2016'))
    dbt = filter_DataTable(_dbt, SCHOOL_DATA['PUPIL_FIELDS'],
            [['SCHOOLYEAR', 'Schuljahr', False]],
            extend = True)
    print("\nINFO:", dbt['__INFO__'])
    print("\nCONTENT:")
    for row in dbt['__ROWS__']:
        print(" :::", row)

    print("\nPUPILS:")
    _dbt = read_DataTable(DATAPATH('testing/delta_test_pupils_2016'))
    dbt = filter_DataTable(_dbt, SCHOOL_DATA['PUPIL_FIELDS'],
            [['SCHOOLYEAR', 'Schuljahr', False]],
            extend = False)
    print("\nINFO:", dbt['__INFO__'])
    print("\nCONTENT:")
    for row in dbt['__ROWS__']:
        print(" :::", row)

    ftype = 'tsv'
    fbytes = make_DataTable(dbt, ftype,
            fieldlist = None, infolist = None, extend = False)
    fpath = DATAPATH('testing/tmp/extended_no') + '.' + ftype
    with open(fpath, 'wb') as fh:
        fh.write(fbytes)
    print("\nSAVED AS:", fpath)

    ftype = 'xlsx'
    fbytes = make_DataTable(dbt, ftype,
            fieldlist = SCHOOL_DATA['PUPIL_FIELDS'],
            infolist = [['SCHOOLYEAR', 'Schuljahr', True]],
            extend = True)
    fpath = DATAPATH('testing/tmp/extended_yes') + '.' + ftype
    with open(fpath, 'wb') as fh:
        fh.write(fbytes)
    print("\nSAVED AS:", fpath)
