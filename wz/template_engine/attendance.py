# -*- coding: utf-8 -*-
"""
template_engine/attendance.py - last updated 2021-12-07

Create attendance table for a class.

==============================
Copyright 2021 Michael Towers

   Licensed under the Apache License, Version 2.0 (the "License");
   you may not use this file except in compliance with the License.
   You may obtain a copy of the License at

       http://www.apache.org/licenses/LICENSE-2.0

   Unless required by applicable law or agreed to in writing, software
   distributed under the License is distributed on an "AS IS" BASIS,
   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
   See the License for the specific language governing permissions and
   limitations under the License.
"""

#TODO: Remake tables with changed pupils, reading data from existing file.

COLOUR_WEEKEND = "DDDDDD"
COLOUR_HOLIDAY = "00CC00"
COLOUR_NO_DAY = "99FFFF"

_PAGE_TITLE = "Klasse {klass} – {month}"

### Messages

_BAD_TEMPLATE_MULTI_TOP_ROW = "ungültige Vorlage: Mehrfach '***'" \
        " in Spalte 'A'\n  {path}"
_BAD_TEMPLATE_NO_TOP_ROW = "ungültige Vorlage: Keine Kopfzeile" \
        " ('***' in Spalte 'A')\n  {path}"
_BAD_TEMPLATE_EXCESS_LINES = "ungültige Vorlage: Nach der letzten Zeile" \
        " mit '*' in Spalte 'A' gibt es weitere Zeilen\n  {path}"
_BAD_TEMPLATE_LINE_TAGS = "ungültige Vorlage: Markierungen in Spalte 'A'" \
        " fehlen\n  {path}"
_BAD_TEMPLATE_HEADER_EMPTY = "ungültige Vorlage: Lücke in Kopfzeile" \
        "\n  {path}"
_BAD_TEMPLATE_EXCESS_COLS = "ungültige Vorlage: überflüssige Spalten" \
        "\n  {path}\nNeue Vorlage erstellt:\n  {path2}"
_BAD_TEMPLATE_HEADER_MISSING = "ungültige Vorlage: Spalte '{h}' fehlt" \
        "\n  {path}"
_TEMPLATE_NEED_MORE_LINES = "die Vorlage braucht mehr Zeilen"\
        " ({n} Schüler)\n  {path}"
_BAD_DATE = "Ungültiges Datum: {key}: {val}"
_BAD_RANGE = "Ungültiges Datumsbereich: {key}: {val}"
_DATE_NOT_IN_YEAR = "Datum liegt nicht im aktuellen Schuljahr: {key}: {val}"
_BAD_SHEET_HEADER = "Ungültiger \"Monatsschlüssel\" (Spalte A) im"\
        " Monat {sheet}\n  Datei: {path}"
_PID_COLUMN_NONZERO = "Tabellendatei {path}:\n  Schülernummer nicht" \
        " in erster Spalte"
_NAME_MISMATCH = "Schüler hat unterschiedliche Namen: {name1}" \
        "und {name2} in\n  {path}"
_CLASS_MISMATCH = "Klasse hat unterschiedliche Namen: {class1}" \
        "und {class2} in\n  {path}"

########################################################################

import sys, os
if __name__ == '__main__':
    import locale
    print("LOCALE:", locale.setlocale(locale.LC_ALL, ''))
    # Enable package import if running as module
    this = sys.path[0]
    appdir = os.path.dirname(this)
    sys.path[0] = appdir
    basedir = os.path.dirname(appdir)
    from core.base import start
#    start.setup(os.path.join(basedir, 'TESTDATA'))
    start.setup(os.path.join(basedir, 'DATA'))

### +++++

from core.base import Dates

import datetime, calendar

from openpyxl import load_workbook
#from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.utils import column_index_from_string#, get_column_letter
from openpyxl.styles import PatternFill

from core.pupils import Pupils
from tables.spreadsheet import Spreadsheet

class AttendanceError(Exception):
    pass

### -----

def month2year(schoolyear, month1, month):
    """Return the year of month <month>. <month1> is the first month in
    the school-year. All values are integers.
    """
    return schoolyear if (month < month1 or month1 == 1) \
            else schoolyear - 1


def get_month_data(calendar_data):
    """Compile the needed calendar data for the given calendar.

    Parameters:
    -----------
    calendar_data (dict):    The contents of a school-year calendar
        config file.

    Returns:
    --------
    month_colours (list):   [
            (month, year, month-name,
                [date-list],
                [colour-list]
            ), ...
        ]
        The month is an <int>, year is the year of that month (<int>),
        month-name is a <str>, the full name.
        date-list and colour-list each have 31 entries (one for each day),
        even when the month has fewer than 31 days. The months start with
        the first month of the school-year.
    """
    hols = read_hols(calendar_data)
    schoolyear = int(Dates.calendar_year(calendar_data))
    month1 = int(calendar_data["YEAR_BEGIN"].split("-")[1])
    #print("??? year, month1:", repr(schoolyear), repr(month1))
    month_colours = []
    month = month1
    while True:
        year = month2year(schoolyear, month1, month)
        col_colour = []
        dates = []
        for day in range(1, 32):
            try:
                date = f"{year}-{month:02}-{day:02}"
                dates.append(date)
                if datetime.date.fromisoformat(date).weekday() > 4:
                    # Weekend
#TODO: What about school Saturdays?
                    col_colour.append(COLOUR_WEEKEND)
                elif date in hols:
                    # Holiday weekday
                    col_colour.append(COLOUR_HOLIDAY)
                else:
                    # A normal schoolday
                    col_colour.append(None)
            except ValueError:
                # date out of range – grey out cell.
                dates.append(None)
                col_colour.append(COLOUR_NO_DAY)
        month_colours.append(
            (   month, year, calendar.month_name[month],
                col_colour,
                dates
            )
        )
        month =(month % 12) + 1
        if month == month1:
            break
    #print("§§§\n", month_colours)
    return month_colours


def read_hols(calendar_data):
    """Return a set of dates for all holidays in the calendar file for
    the current year. The dates are in isoformat (YYYY-MM-DD).

    Parameters:
    -----------
    calendar_data (dict):    The contents of a school-year calendar
        config file.

    Returns:
    --------
        School-free days (dict): {date (isoformat): tag in calendar file}
    """
    schoolyear = Dates.calendar_year(calendar_data)
    deltaday = datetime.timedelta(days = 1)
    hols = {}
    for k, v in calendar_data.items():
        if k[0] == '_':
            if type(v) == str:
                # single date
                try:
                    d = datetime.date.fromisoformat(v)
                except ValueError:
                    raise AttendanceError(_BAD_DATE.format(
                            key=k, val=v))
                if not Dates.check_schoolyear(schoolyear, d=v):
                    raise AttendanceError(_DATE_NOT_IN_YEAR.format(
                            key=k, val=v))
                hols[v] = k
            else:
                d1, d2 = map(datetime.date.fromisoformat, v)
                if d1 >= d2:
                    raise AttendanceError(_BAD_RANGE.format(
                            key=k, val=v))
                while True:
                    hols[d1.isoformat()] = k
                    d1 += deltaday
                    if d1 > d2:
                        break
    return hols


class Table:
    """openpyxl based spreadsheet template handler ('.xlsx'-file) for
    attendance tables.
    """
    def reload_template(self):
        self.workbook = load_workbook(self.template)
        self.sheet0 = self.workbook.active

    def __init__(self):
        self.template = RESOURCEPATH(CONFIG["CLASS_ATTENDANCE_TEMPLATE"])
        self.reload_template()

        # Check formatting
        self.row0 = None    # header line
        self.row1 = None    # first pupil line
        self.rowN = None    # last pupil line
        for cell in self.sheet0["A"]:
            if cell.value == "***":
                if self.row0 is None:
                    self.row0 = cell.row
                else:
                    raise AttendanceError(_BAD_TEMPLATE_MULTI_TOP_ROW.format(
                            path=self.template))
            elif cell.value == "*":
                if self.row0 is None:
                    raise AttendanceError(_BAD_TEMPLATE_NO_TOP_ROW.format(
                            path=self.template))
                if self.row1 is None:
                    self.row1 = cell.row
                self.rowN = cell.row
            elif self.rowN:
                raise AttendanceError(_BAD_TEMPLATE_EXCESS_LINES.format(
                        path=self.template))
        if self.rowN is None:
            raise AttendanceError(_BAD_TEMPLATE_LINE_TAGS.format(
                    path=self.template))
        self.columns = {}   # {header: column-letter}
        col = 0
        colN = None
        for cell in self.sheet0[self.row0]:
            col += 1
            if cell.value:
                self.columns[cell.value] = cell.column_letter
            elif colN is None:
                colN = col
        if colN is not None:
            if len(self.columns) >= colN:
                raise AttendanceError(_BAD_TEMPLATE_HEADER_EMPTY.format(
                    path=self.template))
            else:
                file2 = self.template.rsplit(".xlsx", 1)[0] + "_mod.xlsx"
                self.sheet0.delete_cols(colN, col - colN + 1)
                self.save(file2)
                raise AttendanceError(_BAD_TEMPLATE_EXCESS_COLS.format(
                    path=self.template, path2=file2))
        headers = [f"{d:02}" for d in range(1, 32)] + ["TITLE"]
        for h in headers:
            if h not in self.columns:
                raise AttendanceError(_BAD_TEMPLATE_HEADER_MISSING.format(
                    h=h, path=self.template))

    def getSheetNameFromIndex(self, ix):
        return self.workbook.sheetnames[ix]

    def save(self, filepath):
        self.workbook.save(filepath)

    def make_year(self, schoolyear, klass,
            pupilnames, pupilmap, month_colours):
        """Build a set of attendance tables for the given school-year.

        Parameters:
        -----------
        schoolyear (int):   Year in which the school-year ends
        klass (str):        Class name
        pupilnames (list):  [(pid, pupil-name), ...]
        pupilmap (dict):    {pid: {date: cell-text}, ...}
            dates are in isoformat (YYYY-MM-DD)
        month_colours (list):   [
                (   month, year, month-name,
                    [date-list],
                    [colour-list]
                ), ...
            ]
            The month is an <int>, year is the year of that month (<int>),
            month-name is a <str>, the full name.
            date-list and colour-list each have 31 entries (one for each day),
            even when the month has fewer than 31 days. The months start with
            the first month of the school-year.

        Returns:
        --------
        None
        """
        if self.sheet0 is None:
            self.reload_template()
        self.schoolyear = schoolyear
        self.month1 = month_colours[0][0]
#        self.klass = klass
        self.pupilnames = pupilnames
        self.pupilmap = pupilmap
#        self.month_colours = month_colours

        # First enter the pupils
        row = self.row1
        col_title = self.columns["TITLE"]
        for pid, pname in pupilnames:
            if row > self.rowN:
                raise AttendanceError(_TEMPLATE_NEED_MORE_LINES.format(
                    n=len(pupilnames), path=self.template))
            self.sheet0[f"A{row}"].value = pid
            self.sheet0[f"{col_title}{row}"].value = pname
            row += 1
        if row <= self.rowN:
            self.sheet0.delete_rows(row, self.rowN - row + 1)
        colours = {}
        for month, year, month_name, col_colour, dates in month_colours:
            ws = self.workbook.copy_worksheet(self.sheet0)
            ws.title = month_name
            # Tag the first column for machine-reading:
            ws[f"A{self.row0}"].value = f"#{year}-{month:02}_{klass}"
            # The header for human-reading:
            ws[f"{col_title}{self.row0}"].value = _PAGE_TITLE.format(
                    klass=klass, month=f"{month_name} {year}")
            row = self.row1
            for pid, pname in pupilnames:
                pmap = pupilmap.get(pid) or {}
                for c in range(31):
                    day = f"{c+1:02}"
                    cell = ws[f"{self.columns[day]}{row}"]
                    date = dates[c]
                    colour = col_colour[c]
                    if colour:
                        try:
                            pf = colours[colour]
                        except KeyError:
                            pf = PatternFill("solid", fgColor=colour)
                            colours[colour] = pf
                        cell.fill = pf
                    try:
                        cell.value = pmap[date]
                    except KeyError:
                        pass
                row += 1
        self.workbook.remove(self.sheet0)
        self.sheet0 = None

    def update_table(self, filepath):
        """Take the data from the given attendance file and regenerate
        the file using the pupil data from the database. This allows
        new pupils to be added easily, but also name changes are possible.
        Pupils should probably not be removed, however, as the table is
        supposed to be a record of the whole year.
        """
        last_year_month, klass, pupildata = attendanceTable(filepath,
                self.columns)
#TODO ...


class PupilInfo(Pupils):
    def pupil_list(self, klass):
        pdatalist = self.class_pupils(klass)
        return [(pdata["PID"], pdata.name()) for pdata in pdatalist]


def attendanceTable(filepath, columnmap):
    """Read the file at <filepath> as an attendance table.
    The columns are passed as a mapping: {column-header: column-letter}
    (the letters being those of the columns in a spreadsheet).
    """
    columns = []
    for hdr, col in columnmap.items():
        i = column_index_from_string(col) - 1
        if i == 0:
            if hdr != "***":
                raise AttendanceError(_PID_COLUMN_NONZERO.format(
                        path=filepath))
        else:
            columns.append((hdr, i))
    table = Spreadsheet(filepath)
    pupildata = {}
    klass = None
    last_year_month = ""
    for sheetname in table.getTableNames():
        table.setTable(sheetname)
        sheet = table.table()
        rows = []
        year_month = None
        for row in sheet:
            c1 = row[0]
            if not c1:
                continue
            if year_month:
                # The header line has already been found.
                pid = row[0]
                try:
                    pdata = pupildata[pid]
                except KeyError:
                    pdata = {}
                    pupildata[pid] = pdata
                for f, i in columns:
                    v = row[i]
                    if v:
                        if f == "TITLE":
                            name = pdata.get("__NAME__")
                            if name:
                                if v != name:
                                    raise AttendanceError(
                                        _NAME_MISMATCH.format(
                                            name1=name, name2=v,
                                            path=filepath))
                            else:
                                pdata["__NAME__"] = v
                        else:
                            key = f"{year_month}-{f}"
                            try:
                                pupildata[pid][key] = v
                            except KeyError:
                                pupildata[pid] = {key: v}
            else:
                if c1[0] == "#":
                    try:
                        year_month, k = c1[1:].split("_")
                    except ValueError:
                        raise AttendanceError(_BAD_SHEET_HEADER.format(
                                sheet=name, path = filepath))
                    if klass:
                        if k != klass:
                            raise AttendanceError(_CLASS_MISMATCH.format(
                                class1=klass, class2=k, path = filepath))
                    else:
                        klass = k
                    if year_month > last_year_month:
                        last_year_month = year_month
    return last_year_month, klass, pupildata


#--#--#--#--#--#--#--#--#--#--#--#--#--#--#--#--#--#--#--#--#--#--#--#

if __name__ == '__main__':
    table = Table()
    month_data = get_month_data(CALENDAR)
    pupils = PupilInfo()
    klass = "11G"
    pupilnames = pupils.pupil_list(klass)

#    pupilnames = [
#        ("001", "Fritz Blume"),
#        ("002", "Melanie Dreher"),
#        ("003", "Amelie Lennart"),
#    ]
    table.make_year(int(SCHOOLYEAR), klass,
            pupilnames, {}, month_data)
    outfile = DATAPATH(f"testing/tmp/Anwesenheit_{SCHOOLYEAR}_{klass}.xlsx")
    table.save(outfile)
    print("SAVED TO", outfile)

    infile = DATAPATH("testing/Anwesenheit_test.xlsx")
    indata = attendanceTable(infile, table.columns)
    print(f"\nINPUT FROM {infile}:")
    print(indata)

    quit(0)
    datatable = read_DataTable(outfile)
    print("\n -------------------------------\n")
    print("FIELDS:", datatable["__FIELDS__"])
    i = 0
    for r in datatable["__ROWS__"]:
        i += 1
        print(f"\n{i:02}:", r)
